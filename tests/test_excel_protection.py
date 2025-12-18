import tempfile

from openpyxl import load_workbook

from config import AppConfig, Paths
from core import db, excel, sla

STATUS_COL = "Статус"
COMMENT_COL = "Комментарий"
SUBJECT_COL = "Тема"


def _temp_cfg():
    tmp = tempfile.mkdtemp()
    paths = Paths(
        appdata_dir=tmp,
        log_dir=f"{tmp}/logs",
        db_path=f"{tmp}/test.sqlite3",
        excel_path=f"{tmp}/tickets.xlsx",
        backup_dir=f"{tmp}/backups",
    )
    return AppConfig(paths=paths)


def test_excel_protection_unlocks_editable_columns():
    cfg = _temp_cfg()
    cfg.excel_password = "testpwd"
    db.ensure_schema(cfg)
    conn = db.connect(cfg.paths.db_path, wal_mode=cfg.wal_mode)
    db.seed_test_ticket(conn, status=sla.STATUS_ASSIGNED)
    conn.close()

    path = excel.export_excel(cfg, today_only=False)
    wb = load_workbook(path)
    ws = wb[excel.SHEET_TICKETS]
    header = [c.value for c in ws[1]]
    idx_status = header.index(STATUS_COL) + 1
    idx_comment = header.index(COMMENT_COL) + 1
    idx_subject = header.index(SUBJECT_COL) + 1

    assert ws.protection.sheet is True
    assert ws.cell(row=2, column=idx_status).protection.locked is False
    assert ws.cell(row=2, column=idx_comment).protection.locked is False
    assert ws.cell(row=2, column=idx_subject).protection.locked is True
