# -*- coding: utf-8 -*-
import re
import time
from datetime import datetime, timedelta

import pythoncom
import win32com.client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ===================== НАСТРОЙКИ =====================
DAYS_LOOKBACK =  35         # Сколько дней назад смотреть
EXCEL_FILENAME = "emails.xlsx"
WRAP_COLS = ['B', 'E']        # в Excel: B=Текст, E=Ответ
# =====================================================

def dt_to_restrict_str(dt: datetime) -> str:
    # Outlook Restrict любит формат US 12-часовой: mm/dd/yyyy hh:mm AM/PM
    return dt.strftime("%m/%d/%Y %I:%M %p")

def safe_get(obj, attr, default=None):
    try:
        return getattr(obj, attr)
    except Exception:
        return default

def safe_call(fn, default=None, retries=2, delay=0.05):
    for attempt in range(retries + 1):
        try:
            return fn()
        except pythoncom.com_error:
            if attempt < retries:
                time.sleep(delay * (attempt + 1))
            else:
                return default
        except Exception:
            return default

def clean_text(text: str) -> str:
    if not text:
        return ""
    # один перенос вместо пачки
    text = re.sub(r'\n+', '\n', text)
    # триммим строки
    lines = [ln.strip() for ln in text.split('\n')]
    text = "\n".join(lines)

    # Маркеры подписи: отсекаем только если строка НАЧИНАЕТСЯ с маркера,
    # чтобы не резать нормальные «спасибо» в середине письма
    signature_markers = [
        "с уважением", "с наилучшими пожеланиями", "с уважением,",
        "с наилучшими", "best regards", "regards", "отправлено с",
        "sent from my", "отправлено с моего", "отправлено с моего iphone",
        "sent from my iphone"
    ]
    lower_lines = [ln.lower() for ln in lines]
    cut_at = None
    for i, ln in enumerate(lower_lines):
        if any(ln.startswith(m) for m in signature_markers):
            cut_at = i
            break
        # типичная черта подписи
        if ln in ("--", "—", "––", "___", "_____"):
            cut_at = i
            break
    if cut_at is not None:
        text = "\n".join(lines[:cut_at]).strip()
    return text

def normalize_subject(subject: str) -> str:
    if not subject:
        return ""
    return re.sub(r'^((fw:|fwd:|re:)\s*)+', '', subject, flags=re.IGNORECASE).strip()

def get_sender_smtp(mail) -> str:
    """Надёжно достаём SMTP отправителя."""
    # 1) Пробуем через Sender -> ExchangeUser/Address
    sender = safe_get(mail, "Sender")
    if sender:
        # Exchange
        exch_user = safe_call(lambda: sender.GetExchangeUser(), None)
        if exch_user:
            smtp = safe_get(exch_user, "PrimarySmtpAddress")
            if smtp:
                return smtp
        # Иногда у AddressEntry есть SMTP напрямую
        smtp = safe_get(sender, "Address")
        if smtp:
            return smtp
    # 2) Фолбэк: стандартное свойство
    addr = safe_get(mail, "SenderEmailAddress", "")
    return addr or ""

def iter_folder_items_since(folder, date_field: str, since_dt: datetime):
    """Перебор элементов папки с Restrict/Sort и безопасным курсором."""
    items = folder.Items
    # Сортируем по дате по убыванию (свежие первыми)
    items.Sort(f"[{date_field}]", True)
    # Фильтр по дате
    filt = f"[{date_field}] >= '{dt_to_restrict_str(since_dt)}'"
    restricted = items.Restrict(filt)
    # Перебор через курсор
    item = safe_call(lambda: restricted.GetFirst())
    while item:
        yield item
        item = safe_call(lambda: restricted.GetNext())

def build_sent_index(sent_folder, since_dt: datetime):
    """
    Индекс отправленных:
    key: (ConversationID или нормализованная тема)
    value: {
      'replies':  [(sent_on, body)],
      'forwards': [(sent_on, first_to)]
    } — списки отсортированы по дате (возрастание)
    """
    idx = {}
    for it in iter_folder_items_since(sent_folder, "SentOn", since_dt):
        # Только письма
        cls = safe_get(it, "Class", 0)
        if cls != 43:
            continue

        subject = safe_get(it, "Subject", "") or ""
        sent_on = safe_get(it, "SentOn")
        if not isinstance(sent_on, datetime):
            continue

        conv_id = safe_get(it, "ConversationID")
        # запасной ключ — нормализованная тема
        norm_subj = normalize_subject(subject)
        key = conv_id or norm_subj.lower()

        rec = idx.setdefault(key, {"replies": [], "forwards": []})

        subj_l = subject.strip().lower()
        # тело берём лениво/безопасно
        body = clean_text(safe_get(it, "Body", "") or "")

        if subj_l.startswith("re:"):
            rec["replies"].append((sent_on, body))
        elif subj_l.startswith("fw:") or subj_l.startswith("fwd:"):
            # первый адресат пересылки
            to = safe_get(it, "To", "") or ""
            fwd_to = to.split(";")[0].strip() if to else ""
            rec["forwards"].append((sent_on, fwd_to))

    # Сортируем списки по дате, чтобы потом быстро искать «после received»
    for rec in idx.values():
        rec["replies"].sort(key=lambda x: x[0])
        rec["forwards"].sort(key=lambda x: x[0])
    return idx

def find_first_after(events, after_dt: datetime):
    """events — список (dt, payload) отсортированный по дате; вернуть первый >= after_dt."""
    if not events:
        return None, None
    # Бинарный поиск
    lo, hi = 0, len(events)
    while lo < hi:
        mid = (lo + hi) // 2
        if events[mid][0] < after_dt:
            lo = mid + 1
        else:
            hi = mid
    if lo < len(events):
        return events[lo]
    return None, None

def main():
    pythoncom.CoInitialize()

    outlook = win32com.client.gencache.EnsureDispatch("Outlook.Application")
    ns = outlook.GetNamespace("MAPI")
    inbox = ns.GetDefaultFolder(6)  # Inbox
    sent = ns.GetDefaultFolder(5)   # Sent Items

    since = datetime.now() - timedelta(days=DAYS_LOOKBACK)

    print("Строю индекс отправленных...")
    sent_index = build_sent_index(sent, since)
    print(f"Индекс готов. Ключей: {len(sent_index)}")

    rows = []
    processed = 0
    started = time.time()

    print("Обрабатываю входящие...")
    for it in iter_folder_items_since(inbox, "ReceivedTime", since):
        try:
            cls = safe_get(it, "Class", 0)
            if cls != 43:  # только MailItem
                continue

            received = safe_get(it, "ReceivedTime")
            if not isinstance(received, datetime):
                continue

            subject = safe_get(it, "Subject", "") or ""
            body_raw = safe_get(it, "Body", "") or ""
            body = clean_text(body_raw)
            sender = get_sender_smtp(it)

            conv_id = safe_get(it, "ConversationID")
            norm_subj = normalize_subject(subject)
            key = (conv_id or norm_subj.lower())

            reply_date_str = ""
            reply_body = ""
            forward_recipient = ""

            rec = sent_index.get(key)
            if rec:
                # Ищем первый ответ/пересылку ПОСЛЕ получения
                rdt, rbody = find_first_after(rec["replies"], received)
                fdt, fto   = find_first_after(rec["forwards"], received)
                # Приоритет: если есть и ответ и пересылка — оставим оба (как в твоей логике — статус пуст)
                if rdt:
                    reply_date_str = rdt.strftime("%d.%m.%Y")
                    reply_body = clean_text(rbody or "")
                if fdt and fto:
                    forward_recipient = fto

            status = "В работе" if (not reply_date_str and not forward_recipient) else ""

            rows.append({
                "Дата": received.strftime("%d.%m.%Y"),
                "Текст": body,
                "Перенаправление": forward_recipient,
                "Дата ответа": reply_date_str,
                "Ответ": reply_body,
                "Статус": status,
                "Тема письма": subject,
                "От кого": sender
            })

            processed += 1
            if processed % 50 == 0:
                print(f"Входящих обработано: {processed} за {time.time() - started:.1f}s")

        except Exception as e:
            # Логгируем, но не падаем на «битых» письмах
            print(f"Ошибка на входящем: {e}")
            continue

    df = pd.DataFrame(rows)
    if df.empty:
        print("Нет данных для записи в Excel.")
        return

    # Порядок столбцов
    df = df[["Дата", "Текст", "Перенаправление", "Дата ответа", "Ответ", "Статус", "Тема письма", "От кого"]]
    df.to_excel(EXCEL_FILENAME, index=False)
    print(f"Данные сохранены в {EXCEL_FILENAME}")

    # Лёгкое форматирование
    wb = load_workbook(EXCEL_FILENAME)
    ws = wb.active
    for col in WRAP_COLS:
        ws.column_dimensions[col].width = 50
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True)
    wb.save(EXCEL_FILENAME)
    print("Форматирование Excel завершено.")

if __name__ == "__main__":
    main()
