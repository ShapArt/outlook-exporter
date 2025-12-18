from __future__ import annotations

import os
import shutil
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from config import AppConfig
from core import db
from core.logger import get_logger
from core.sla import (
    STATUS_MODEL,
    normalize_status_code,
    status_code_to_label,
    status_text_to_code,
)

log = get_logger(__name__)

SHEET_TICKETS = "Заявки"
SHEET_OVERDUE = "Просрочки"
SHEET_KPI = "KPI"
SHEET_CONFLICTS = "Конфликты"
SHEET_STATUS_HELP = "Статусы и подсказки"

STATUS_COLORS: Dict[str, str] = {
    "resolved": "C6EFCE",
    "responded": "C6EFCE",
    "assigned": "FFF2CC",
    "new": "FFF2CC",
    "overdue": "F8CBAD",
    "not_interesting": "D9D9D9",
    "waiting_customer": "D9E1F2",
    "table": "D9E1F2",
    "otip": "FCE4D6",
}

PRIORITY_LIST = ["p1", "p2", "p3", "p4"]
DEFAULT_STATUS_LIST = [code for code, _ in STATUS_MODEL]

COLUMN_WIDTHS = {
    "ticket_id": 10,
    "stable_id": 12,
    "row_version": 10,
    "ID заявки": 12,
    "Получено": 18,
    "Тема": 40,
    "Отправитель (SMTP)": 28,
    "Email клиента": 28,
    "Текст заявки": 42,
    "Статус": 18,
    "Ответственный": 24,
    "Первый ответ": 18,
    "Текст ответа": 28,
    "Переслано кому": 22,
    "SLA дедлайн": 18,
    "Просрочка": 12,
    "Комментарий": 40,
    "Рекомендованный ответ": 40,
    "Повторы (30д)": 24,
    "Приоритет": 12,
    "Тема/группа": 24,
    "Обновлено": 18,
    "Кем обновлено": 18,
    "Источник": 14,
}
DEFAULT_COL_WIDTH = 18

WRAP_COLUMNS = {
    "Тема",
    "Текст заявки",
    "Комментарий",
    "Рекомендованный ответ",
    "Повторы (30д)",
    "Текст ответа",
    "Тема/группа",
}

RUS_COLUMNS = [
    ("ticket_id", "ticket_id"),
    ("stable_id", "stable_id"),
    ("row_version", "row_version"),
    ("id_display", "ID заявки"),
    ("first_received_utc", "Получено"),
    ("subject", "Тема"),
    ("sender", "Отправитель (SMTP)"),
    ("customer_email", "Email клиента"),
    ("body", "Текст заявки"),
    ("status", "Статус"),
    ("responsible", "Ответственный"),
    ("first_reply_utc", "Первый ответ"),
    ("first_reply_body", "Текст ответа"),
    ("first_forward_to", "Переслано кому"),
    ("sla_due", "SLA дедлайн"),
    ("overdue", "Просрочка"),
    ("comment", "Комментарий"),
    ("recommended_answer", "Рекомендованный ответ"),
    ("repeat_hint", "Повторы (30д)"),
    ("priority", "Приоритет"),
    ("topic", "Тема/группа"),
    ("last_updated_at", "Обновлено"),
    ("last_updated_by", "Кем обновлено"),
    ("data_source", "Источник"),
]

# Backward compatibility: recognise old headers if user edits legacy Excel
LEGACY_COL_ALIASES = {
    "Статус": ["ö‘'ø‘'‘?‘?", "ÿ?ö'ÿ‚øö'ö?ö?"],
    "Ответственный": ["?‘'?ç‘'‘?‘'?ç??‘<ü", "ÿ?ö'ÿ?ÿ‘Øö'ö?ö'ÿ?ÿ‘Øÿ?ÿ?ö<ÿ¢\"-"],
    "Комментарий": ["????ç?‘'ø‘?ñü", "ÿ?ÿ?ÿ?ÿ?ÿ‘Øÿ?ö'ÿ‚øö?ÿ‘'ÿ¢\"-"],
    "Приоритет": ["?‘?ñ?‘?ñ‘'ç‘'", "ÿ?ö?ÿ‘'ÿ?ö?ÿ‘'ö'ÿ‘Øö'"],
    "Текст заявки": [
        "÷çó‘?‘' ?+‘?ø‘%ç?ñ‘?",
        "ÿ‘?ÿ‚çÿ‘\"ö'ö?ö' ÿ‘ÿ‚+ö'ÿ‚øö¢?øÿ‚çÿ:ÿ‘'ö?",
        "ÿÅÿ‘ÿ‘-ö'ÿ‘ö?",
    ],
}


def _header(df: pd.DataFrame, name: str) -> Optional[str]:
    if name in df.columns:
        return name
    for alias in LEGACY_COL_ALIASES.get(name, []):
        if alias in df.columns:
            return alias
    return None


def _resolve_excel_password(cfg: AppConfig) -> str:
    return os.environ.get("NAOS_EXCEL_PASSWORD", cfg.excel_password or "naos")


def export_excel(
    cfg: AppConfig, today_only: bool = False, conflicts: Optional[List[Dict]] = None
) -> Path:
    db.ensure_schema(cfg)
    conn = db.connect(cfg.paths.db_path, wal_mode=cfg.wal_mode)
    where = ""
    params: Tuple = ()
    if today_only:
        where = "date(first_received_utc)=date('now','localtime')"
    rows = db.fetch_tickets(conn, where, params)
    conn.close()

    if not rows:
        log.info("No tickets to export, creating empty workbook with headers")
        df = pd.DataFrame(columns=[col for _, col in RUS_COLUMNS])
    else:
        records = []
        for r in rows:
            received = r["first_received_utc"]
            due = None
            if r["priority"]:
                try:
                    prio_cfg = cfg.sla_by_priority.get(r["priority"], {})
                    hours = float(
                        prio_cfg.get("resolution_hours", cfg.overdue_days * 24)
                    )
                    due = datetime.fromisoformat(received) + timedelta(hours=hours)
                except Exception:
                    due = None
            records.append(
                {
                    "ticket_id": r["id"],
                    "stable_id": r["stable_id"],
                    "row_version": r["row_version"],
                    "id_display": r["id"],
                    "first_received_utc": received,
                    "subject": r["subject"],
                    "sender": r["sender"],
                    "customer_email": r["customer_email"] or "",
                    "body": r["body"],
                    "status": status_code_to_label(r["status"]),
                    "responsible": r["responsible"] or "",
                    "first_reply_utc": r["first_reply_utc"],
                    "first_reply_body": r["first_reply_body"],
                    "first_forward_to": r["first_forward_to"],
                    "sla_due": due.isoformat() if due else "",
                    "overdue": "Да" if r["overdue"] else "Нет",
                    "comment": r["comment"] or "",
                    "recommended_answer": r["recommended_answer"] or "",
                    "repeat_hint": r["repeat_hint"] or "",
                    "priority": r["priority"] or "",
                    "topic": r["topic"] or "",
                    "last_updated_at": r["last_updated_at"],
                    "last_updated_by": r["last_updated_by"],
                    "data_source": r["data_source"],
                }
            )
        df = pd.DataFrame(records, columns=[key for key, _ in RUS_COLUMNS])
        df.columns = [col for _, col in RUS_COLUMNS]

    status_list_raw = cfg.status_catalog or DEFAULT_STATUS_LIST
    status_list = [normalize_status_code(s) or s for s in status_list_raw]

    with tempfile.NamedTemporaryFile(
        prefix="tickets_", suffix=".xlsx", delete=False
    ) as tmp:
        tmp_path = Path(tmp.name)
    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=SHEET_TICKETS, index=False)
        overdue_df = (
            df[df["Просрочка"] == "Да"]
            if not df.empty
            else pd.DataFrame(columns=df.columns)
        )
        overdue_df.to_excel(writer, sheet_name=SHEET_OVERDUE, index=False)
        kpi_rows = _build_kpi(df)
        pd.DataFrame(kpi_rows, columns=["Показатель", "Значение"]).to_excel(
            writer, sheet_name=SHEET_KPI, index=False
        )
        conflicts_df = pd.DataFrame(
            conflicts or [], columns=["ticket_id", "поле", "excel", "db", "решение"]
        )
        conflicts_df.to_excel(writer, sheet_name=SHEET_CONFLICTS, index=False)
        hints = []
        for st in status_list:
            hints.append(
                {
                    "Статус": status_code_to_label(st),
                    "Подсказка": cfg.status_hints.get(st, ""),
                }
            )
        pd.DataFrame(hints).to_excel(writer, sheet_name=SHEET_STATUS_HELP, index=False)

    _decorate_excel(
        tmp_path, status_list=status_list, password=_resolve_excel_password(cfg)
    )
    final_path = _atomic_replace(tmp_path, cfg.paths.excel_path)
    log.info("Excel exported to %s", final_path)
    return final_path


def _decorate_excel(
    path: Path, status_list: Optional[List[str]] = None, password: str = "naos"
) -> None:
    status_list = [
        normalize_status_code(s) or s for s in (status_list or DEFAULT_STATUS_LIST)
    ]
    wb = load_workbook(path)
    editable_headers = {"Статус", "Ответственный", "Комментарий", "Приоритет"}
    for sheet_name in (SHEET_TICKETS, SHEET_OVERDUE):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_names = [cell.value for cell in ws[1]]
        for col in ws.columns:
            col_letter = col[0].column_letter
            header_val = col[0].value
            target_width = COLUMN_WIDTHS.get(header_val, DEFAULT_COL_WIDTH)
            ws.column_dimensions[col_letter].width = target_width
            if header_val in WRAP_COLUMNS:
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if "Статус" in header_names:
                status_idx = header_names.index("Статус")
                status_val = (row[status_idx].value or "").strip()
                code = status_text_to_code(status_val)
                color = STATUS_COLORS.get(code or status_val)
                if color:
                    fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                    for cell in row:
                        cell.fill = fill
            for cell in row:
                header = header_names[cell.column - 1]
                cell.protection = Protection(locked=header not in editable_headers)

        for col_name in ("ticket_id", "stable_id", "row_version"):
            if col_name in header_names:
                idx = header_names.index(col_name)
                col_letter = ws[1][idx].column_letter
                ws.column_dimensions[col_letter].hidden = True

        if "Статус" in header_names:
            status_idx = header_names.index("Статус") + 1
            labels = [status_code_to_label(s) for s in status_list]
            dv_status = DataValidation(
                type="list", formula1=f'"{",".join(labels)}"', allow_blank=True
            )
            dv_status.errorTitle = "Неверный статус"
            dv_status.error = "Выберите статус из выпадающего списка."
            ws.add_data_validation(dv_status)
            dv_status.add(
                f"{ws.cell(row=2, column=status_idx).coordinate}:{ws.cell(row=1000, column=status_idx).coordinate}"
            )
        if "Приоритет" in header_names:
            prio_idx = header_names.index("Приоритет") + 1
            dv_prio = DataValidation(
                type="list", formula1=f'"{",".join(PRIORITY_LIST)}"', allow_blank=True
            )
            dv_prio.errorTitle = "Неверный приоритет"
            dv_prio.error = "Используйте p1/p2/p3/p4."
            ws.add_data_validation(dv_prio)
            dv_prio.add(
                f"{ws.cell(row=2, column=prio_idx).coordinate}:{ws.cell(row=1000, column=prio_idx).coordinate}"
            )

        ws.protection.sheet = True
        ws.protection.password = password or "naos"
        ws.protection.enable()

    if SHEET_KPI in wb.sheetnames:
        ws_kpi = wb[SHEET_KPI]
        ws_kpi.freeze_panes = "A2"
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        for cell in ws_kpi[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws_kpi.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws_kpi.column_dimensions[col_letter].width = min(max_length + 4, 60)
    if SHEET_STATUS_HELP in wb.sheetnames:
        ws_ref = wb[SHEET_STATUS_HELP]
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        for cell in ws_ref[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col in ws_ref.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws_ref.column_dimensions[col_letter].width = min(max_length + 4, 60)

    wb.save(path)


def _atomic_replace(src: Path, dst: Path) -> Path:
    dst.parent.mkdir(parents=True, exist_ok=True)
    backup = None
    if dst.exists():
        backup = dst.with_suffix(".bak")
        try:
            shutil.copy2(dst, backup)
        except Exception as exc:
            log.warning("Backup failed: %s", exc)
    try:
        src.replace(dst)
        return dst
    except PermissionError as exc:
        pending = dst.with_name(dst.stem + "_pending" + dst.suffix)
        try:
            src.replace(pending)
            log.error("Excel file is locked: %s; saved copy to %s", dst, pending)
            return pending
        except Exception:
            log.error("Excel file is locked and pending save failed: %s", dst)
            raise exc
    finally:
        if src.exists():
            try:
                src.unlink()
            except Exception:
                pass


def _read_ticket_sheet(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    candidates = [SHEET_TICKETS, "Tickets"] + list(xls.sheet_names)
    for name in candidates:
        if name in xls.sheet_names:
            try:
                return pd.read_excel(path, sheet_name=name)
            except Exception:
                continue
    raise ValueError("Не удалось найти лист с заявками")


def sync_from_excel(cfg: AppConfig) -> Dict[str, int]:
    db.ensure_schema(cfg)
    path = cfg.paths.excel_path
    if not path.exists():
        log.info("Excel file not found, skip sync")
        return {"updated": 0, "conflicts": 0, "missing": 0}
    try:
        df = _read_ticket_sheet(path)
    except Exception as exc:
        log.error("Failed to read excel: %s", exc)
        return {"updated": 0, "conflicts": 0, "missing": 0}

    status_col = _header(df, "Статус")
    resp_col = _header(df, "Ответственный")
    prio_col = _header(df, "Приоритет")
    comment_col = _header(df, "Комментарий")
    required = {"ticket_id", "row_version", status_col, resp_col}
    if None in required or not required.issubset(set(df.columns)):
        log.error("Excel missing required columns")
        return {"updated": 0, "conflicts": 0, "missing": 0}

    conn = db.connect(cfg.paths.db_path, wal_mode=cfg.wal_mode)
    updated = conflicts = missing = 0
    conflict_rows: List[Dict] = []
    try:
        for _, row in df.iterrows():
            try:
                ticket_id = int(row["ticket_id"])
            except Exception:
                continue
            excel_rv = int(row["row_version"])
            status_display = str(row.get(status_col, "")).strip()
            status = status_text_to_code(status_display) or status_display
            responsible = str(row.get(resp_col, "") or "").strip()
            priority = str(row.get(prio_col, "") or "").strip()
            comment = str(row.get(comment_col, "") or "").strip()
            excel_updated_at = str(row.get("Обновлено", "") or "").strip()
            excel_updated_dt = None
            try:
                excel_updated_dt = datetime.fromisoformat(excel_updated_at)
            except Exception:
                excel_updated_dt = None
            cur = conn.execute(
                "SELECT row_version, last_updated_at FROM tickets WHERE id=?",
                (ticket_id,),
            )
            current = cur.fetchone()
            if not current:
                missing += 1
                continue
            db_updated = current["last_updated_at"]
            db_updated_dt = None
            try:
                db_updated_dt = (
                    datetime.fromisoformat(db_updated) if db_updated else None
                )
            except Exception:
                db_updated_dt = None

            if db_updated_dt and excel_updated_dt and db_updated_dt > excel_updated_dt:
                conflicts += 1
                conflict_rows.append(
                    {
                        "ticket_id": ticket_id,
                        "поле": "last_updated_at",
                        "excel": excel_updated_at,
                        "db": db_updated,
                        "решение": "skip (db newer)",
                    }
                )
                db.log_event(
                    conn,
                    ticket_id,
                    "excel_conflict",
                    None,
                    status,
                    "excel",
                    raw_response="db newer",
                )
                continue
            if int(current["row_version"]) != excel_rv:
                conflicts += 1
                conflict_rows.append(
                    {
                        "ticket_id": ticket_id,
                        "поле": "row_version",
                        "excel": excel_rv,
                        "db": current["row_version"],
                        "решение": "skip",
                    }
                )
                db.log_event(
                    conn,
                    ticket_id,
                    "excel_conflict",
                    None,
                    status,
                    "excel",
                    raw_response="row_version mismatch",
                )
                continue
            conn.execute(
                """
                UPDATE tickets
                SET status=?, responsible=?, priority=?, comment=?, overdue=CASE WHEN ? IN ('resolved','table','in_table','not_interesting') THEN 0 ELSE overdue END,
                    row_version=row_version+1, last_status_utc=datetime('now'), last_updated_at=datetime('now'), last_updated_by='excel', data_source='excel'
                WHERE id=? AND row_version=?
                """,
                (
                    status,
                    responsible or None,
                    priority or None,
                    comment or None,
                    status,
                    ticket_id,
                    excel_rv,
                ),
            )
            db.log_event(conn, ticket_id, "excel_sync", None, status, "excel")
            updated += 1
        conn.commit()
    finally:
        conn.close()
    log.info(
        "Excel sync completed: %s updated, %s conflicts, %s missing",
        updated,
        conflicts,
        missing,
    )
    return {
        "updated": updated,
        "conflicts": conflicts,
        "missing": missing,
        "conflict_rows": conflict_rows,
    }


def _build_kpi(df: pd.DataFrame) -> List[Tuple[str, object]]:
    rows: List[Tuple[str, object]] = []
    total = len(df)
    overdue = int((df["Просрочка"] == "Да").sum()) if not df.empty else 0
    rows.append(("Всего заявок", total))
    rows.append(("Просрочка", overdue))
    if total:
        rows.append(("Доля просрочек", f"{(overdue/total*100):.1f}%"))
    for status in df.get("Статус", pd.Series(dtype=object)).unique():
        if pd.isna(status):
            continue
        cnt = int((df["Статус"] == status).sum())
        rows.append((f"Статус {status}", cnt))
    if "Ответственный" in df.columns and not df.empty:
        resp_counts = df["Ответственный"].fillna("").value_counts().head(5)
        for name, cnt in resp_counts.items():
            label = name or "<не задан>"
            rows.append((f"Топ ответственное лицо: {label}", int(cnt)))
    return rows
